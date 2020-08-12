#! python3
# spreadsheetInverter.py - Invert the row and column of the cells in the
#                          spreadsheet.
# Adam Pellot

import openpyxl
import os

os.chdir('C:\\Users\\Adam\\Documents')

print('Which file would you like to invert?')
filename = input()

# Create a new workbook.
wb = openpyxl.Workbook()
sheet = wb['Sheet']


# Open the requested file.
oldwb = openpyxl.load_workbook(filename)
oldSheet = oldwb[oldwb.sheetnames[0]]

# Copy the values of the old sheet to a 2D list.
rows = oldSheet.max_row + 1
sheetData = [[] for i in range(oldSheet.max_row)]

for rowNum in range(0, oldSheet.max_row, 1):
    for columnNum in range(0, oldSheet.max_column, 1):
        sheetData[rowNum].append(oldSheet.cell(
            row=rowNum + 1, column=columnNum + 1).value)

# Invert the values saved on the 2D list and place them on the new sheet.
for rowNum in range(0, oldSheet.max_column, 1):
    for columnNum in range(0, oldSheet.max_row, 1):
        sheet.cell(row=rowNum + 1,
                   column=columnNum + 1).value = sheetData[columnNum][rowNum]

# Save the new workbook and close the files.
wb.save('new' + filename)
oldwb.close()
wb.close()

print('Done.')
