#! python3
# multiplicationTable.py - Takes a number N from the command line and
#                          creates an N×N multiplication table in an
#                          Excel spreadsheet.
# Adam Pellot

import os
import openpyxl
from openpyxl.styles import Font
import sys

os.chdir('C:\\Users\\Adam\\Documents')

# Declares number n from command line arguments.
inputedNum = sys.argv[1]
n = int(inputedNum)

# Create new workbook.
wb = openpyxl.Workbook()
sheet = wb['Sheet']

# Setup bold font.
boldFont = Font(bold=True)

# Setup the rows of the table.
rowNum = 2
columnNum = 1
for i in range(columnNum, n + 1, 1):
    sheet.cell(row=rowNum, column=columnNum).value = i
    sheet.cell(row=rowNum, column=columnNum).font = boldFont
    rowNum += 1

# Setup the columns of the table.
rowNum = 1
columnNum = 2
for i in range(rowNum, n + 1, 1):
    sheet.cell(row=rowNum, column=columnNum).value = i
    sheet.cell(row=rowNum, column=columnNum).font = boldFont
    columnNum += 1

# Fill table.
rowNum = 2
rowValue = sheet['A2'].value
for i in range(2, n + 2, 1):
    columnNum = 2
    columnValue = sheet['B1'].value
    for i in range(2, n + 2, 1):
        sheet.cell(row=rowNum, column=columnNum).value = rowValue * columnValue
        columnNum += 1
        columnValue += 1
    rowNum += 1
    rowValue += 1

wb.save('multiplicationTableFor' + inputedNum + '.xlsx')
