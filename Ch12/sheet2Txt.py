#! python3
# sheet2Txt.py - Open a spreadsheet and write the cells of column A into
#                one text file, the cells of column B into another text
#                file, and so on.
# Adam Pellot

import os
import openpyxl

# Directory with the spreadsheet.
os.chdir('C:\\Users\\Adam\\Documents')

print('Enter the name of the spreadsheet you would like to use')
filename = input()

# Open the workbook.
wb = openpyxl.load_workbook(filename)
sheet = wb[wb.sheetnames[0]]

# Create the text file names.
fileNames = []
for column in range(sheet.max_column):
    fileNames.append('file' + str(column + 1) + '.txt')

# Write the contents of each column to it's own text file.
print('Writing to text files....')
for column in range(sheet.max_column):
    file = open(fileNames[column], 'w')
    for rowNum in range(sheet.max_row):
        file.write(str(sheet.cell(row=rowNum + 1, column=column + 1).value) +
                   '\n')
    file.close()

# Save the new workbook and close the file.
wb.save('newSpreadSheet.xlsx')
wb.close()

print('Done.')
