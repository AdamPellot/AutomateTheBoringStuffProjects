#! python3
# blankRowInserter.py - Takes two integers and a filename string as
#                       command line arguments. Let’s call the first
#                       integer N and the second integer M. Starting at
#                       row N, the program should insert M blank rows
#                       into the spreadsheet.
# Adam Pellot

import openpyxl
import os
import sys

os.chdir('C:\\Users\\Adam\\Documents')

# Create a new workbook.
wb = openpyxl.Workbook()
sheet = wb['Sheet']

n = sys.argv[1]
n = int(n)
m = sys.argv[2]
m = int(m)
filename = sys.argv[3]

# Open the file passed.
oldwb = openpyxl.load_workbook(filename)
oldSheet = oldwb[oldwb.sheetnames[0]]

# Copy the rows before the row where blank rows are supposed to start.
copiedCells = []
count = 1
stop = n
for row in oldSheet.values:
    for value in row:
        copiedCells.append(value)
    count += 1
    if count == stop:
        break

# Add the copied rows to new sheet.
i = 0
for rowNum in range(1, n, 1):
    columnNum = 1
    for columnNum in range(1, oldSheet.max_column + 1, 1):
        sheet.cell(row=rowNum, column=columnNum).value = copiedCells[i]
        i += 1

# Copy the rest of the rows after the blank rows.
copiedCells = []
for rowNum in range(n, oldSheet.max_row + 1, 1):
    columnNum = 1
    for columnNum in range(1, oldSheet.max_column + 1, 1):
        copiedCells.append(oldSheet.cell(row=rowNum, column=columnNum).value)

# Add the remaining copied rows after m blank rows.
i = 0
for rowNum in range(n + m, oldSheet.max_row + m + 1, 1):
    columnNum = 1
    for columnNum in range(1, oldSheet.max_column + 1, 1):
        sheet.cell(row=rowNum, column=columnNum).value = copiedCells[i]
        i += 1

# Save the new workbook and close the files.
wb.save('new' + filename)
wb.close()
oldwb.close()

print('Done.')
