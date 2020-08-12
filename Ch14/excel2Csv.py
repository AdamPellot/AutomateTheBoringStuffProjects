#! python3
# excel2Csv.py - Reads all the Excel files in the current working
#                directory and outputs them as CSV files
# Adam Pellot

import csv
import os
import openpyxl

print('Enter the absolute path of the directory you want to use:')
directory = input()
os.chdir(directory)

for excelFile in os.listdir('.'):
    # Skip non-xlsx files, load the workbook object.
    if not excelFile.endswith('.xlsx'):
        continue
    else:
        wb = openpyxl.load_workbook(excelFile)
    shtNum = 1
    for sheetName in wb.sheetnames:
        # Loop through every sheet in the workbook.
        sheet = wb[sheetName]

        # Create the CSV filename from the Excel filename and sheet title.
        csvFileObj = open(os.path.join(excelFile[:-5] + 'Sheet' +
                                       str(shtNum) + '.csv'), 'w', newline='')
        # Create the csv.writer object for this CSV file.
        csvWriter = csv.writer(csvFileObj)

        # Loop through every row in the sheet.
        for rowNum in range(1, sheet.max_row + 1):
            rowData = []    # append each cell to this list
            # Loop through each cell in the row.
            for colNum in range(1, sheet.max_column + 1):
                # Append each cell's data to rowData.
                rowData.append(sheet.cell(row=rowNum, column=colNum).value)
            # Write the rowData list to the CSV file.
            csvWriter.writerow(rowData)
        csvFileObj.close()
